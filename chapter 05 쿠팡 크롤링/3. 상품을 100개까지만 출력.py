# 키워드 별로 순위, 브랜드명, 가격, 상세페이지링크 엑셀 저장

# 1~100위까지

#len([])

#게이밍 마우스, 기계식 키보드, 27인치 모니터

import requests
from bs4 import BeautifulSoup
import pyautogui
import openpyxl


#select 는 한 페이지에 여러개
#select_one는 각 페이지의 이름 과 같이

# 헤더에 User-Agent, Accept-Language 를 추가하지 않으면 멈춥니다
header = {
    'Host': 'www.coupang.com',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:76.0) Gecko/20100101 Firefox/76.0',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'ko-KR,ko;q=0.8,en-US;q=0.5,en;q=0.3',
}



# step1+2: 입력부분
keyword = pyautogui.prompt("검색어를 입력하시오:")

##lastpage = int(pyautogui.prompt("몇 페이지까지 크롤링 할까요?"))


# 엑셀 생성하기
# wb = openpyxl.Workbook('coupang_result.xlsx')


# 엑셀 생성하기
wb = openpyxl.Workbook()


# 입력한 키워드 명으로 엑셀 시트 생성
ws = wb.create_sheet(keyword)

# 처음 새로운 엑셀파일이 생성될 때 존재하는 "Sheet" 시트 삭제
wb.remove(wb['Sheet'])

#열 너비 조절
ws.column_dimensions['A'].width = 5.0
ws.column_dimensions['B'].width = 10.0
ws.column_dimensions['C'].width = 70.0
ws.column_dimensions['D'].width = 10.0
ws.column_dimensions['E'].width = 120

# 리스트 형태로 맨위의 칸 추가
ws.append(['순위', '브랜드명', '상품명', '가격', '상세페이지링크'])


#rank: 광고가 아닌 상품의 100 카운트
rank = 1


#이중for문을 빠져나오기 위한 flag
done = False


# step 3 연산부
for page in range(1, 5):
    #이중 반복문 탈출
    if done == True:
        break


    done = False
    print(page, "페이지 입니다=================")
    main_url = f"https://www.coupang.com/np/search?q={keyword}&channel=user&component=&eventCategory=SRP&trcid=&traid=&sorter=scoreDesc&minPrice=&maxPrice=&priceRange=&filterType=&listSize=36&filter=&isPriceRange=false&brand=&offerCondition=&rating=0&page={page}&rocketAll=false&searchIndexingToken=1=6&backgroundColor="

    # headers={'user-agent':'Mozila/5.0'} 봇으로 접근 가능하게 함 headers={'user-agent':'Mozila/5.0'}
    response = requests.get(main_url, headers=header)
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')

    # select의 결과는 리스트 자료형
    links = soup.select("a.search-product-link")


    for link in links:
        # 광고 상품 제거

        if len(link.select("span.ad-badge-text"))>0:
            print("광고 상품 입니다")
        else:
            sub_url = "https://www.coupang.com/" + link.attrs['href']

            # 상세 페이지 링크
            response = requests.get(sub_url, headers=header)
            html = response.text
            soup = BeautifulSoup(html, 'html.parser')

            # 브랜드명 추출     , Text요소만 출력
            # 브랜드명은 있을 수도 없을 수도 있음
            # 중고 상품일 때는 태그가 달라짐
            # try -except로 예외처리
            # 예외가 났을 때 브랜드명 비워줌
            try:
                brand_name = soup.select_one("a.prod-brand-name").text
                # print(brand_name)
            except:
                brand_name = ""

            # 상품명 추출
            name = soup.select_one("h2.prod-buy-header__title").text
            # print(name)

            # 가격 추출
            # 만약 와우 할인가가 있으면 와우 할인가 출력
            # 그렇지 않으면 쿠팡 판매가 출력

            # 쿠팡 판매가
            try:
                price = soup.select_one("span.total-price > strong").text
            except:
                price = 0
            print(rank, brand_name, name, price)
            ws.append([rank, brand_name, name, price, sub_url])
            rank = rank + 1
            if rank > 100:
                done = True
                break


# 엑셀 저장하기
wb.save(f"쿠팡 크롤링 결과/{keyword}.xlsx")

# 특정 이름으로 저장하기
# wb.save('쿠팡 크롤링 결과/coupang_result.xlsx')

## 하이퍼링크를 달기 위한 코드 과정
from openpyxl import load_workbook
wb = load_workbook(f"쿠팡 크롤링 결과/{keyword}.xlsx")

# 시트 활성화
ws = wb.active

# 하이퍼링크 활성화
for i in range(2, ws.max_row+1):
    ws['E' + str(i)].hyperlink = ws['E'+str(i)].value
    ws['E' + str(i)].style = "Hyperlink"
wb.save(f"쿠팡 크롤링 결과/{keyword}_hyperlink.xlsx")



