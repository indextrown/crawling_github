import openpyxl

#1 엑셀 만들기
wb = openpyxl.Workbook()

#2 엑셀 워크시트 만들기
ws = wb.create_sheet('오징어게임')



#3 데이터 추가하기
ws['A1'] = '참가번호'
ws['B1'] = '성명'

ws['A2'] = '1'
ws['B2'] = '오일남'

#4 엑셀 저장하기
wb.save(r"C:\Users\Index\PycharmProjects\startproject\chapter03\엑셀다루기\참가자_data.xlsx")
##문자열 안에서 \는 이스케이프 역할을 하기 때문에
##그냥 문자로 인식하기 위해 r을 입력 (안하면 \\ 또는 /로 해야함)