from openpyxl import Workbook

# 엑셀 생성하기
wb = Workbook()

# 엑셀 시트 생성하기
ws = wb.create_sheet("startcoding")

# 셀 데이터 추가하기
ws['A1'] = "스타트코딩"

# 엑셀 저장하기
wb.save("test.xlsx")
