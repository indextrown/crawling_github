import requests
from bs4 import BeautifulSoup
import time
import pyautogui
from openpyxl import Workbook
from openpyxl.styles import Alignment  #엑셀 자동 줄바꿈




#사용자 입력
keyword = pyautogui.prompt("검색어를 입력하시오:")
#pyautogui.prompt는 리턴 데이터 기본 값은 문자열임 그래서 int로 강제형변환
lastpage = int(pyautogui.prompt("몇 페이지까지 크롤링 할까요?"))

# 엑셀 생성하기
wb = Workbook()

# 엑셀 시트 생성하기
ws = wb.create_sheet(keyword)

#열 너비 조절
ws.column_dimensions['A'].width = 60
ws.column_dimensions['B'].width = 60
ws.column_dimensions['C'].width = 120
#행 번호
row = 1

#페이지 번호
page_num=1

#여러 페이지 가져오기
for i in range(1, lastpage * 10, 10):
    print(f"{page_num}페이지 크롤링 중입니다.===========")
    response = requests.get(
        f"https://search.naver.com/search.naver?where=news&sm=tab_jum&query={keyword}&start={i}")
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')

    # 뉴스 기사 div 10개 추출
    articles = soup.select("div.news_info")

    #
    for article in articles:
        links = article.select("a.info")

        # 링크가 2개 이상이면
        if len(links) >= 2:
            # 두번째 링크의 href를 추출한다
            url = links[1].attrs['href']

            # headers={'user-agent':'Mozila/5.0'} 봇으로 접근 가능하게 함
            response = requests.get(url, headers={'user-agent': 'Mozila/5.0'})
            html = response.text
            soup = BeautifulSoup(html, 'html.parser')

            content = soup.select_one("#dic_area")
            title = soup.select_one(".media_end_head_title")
                # 만약 연애 뉴스 라면 //response => 링크가 리다이렉션인지(사이트 주소가 바뀌는지) 확인하기 위해 response.url 로 만듬

            # 만약 연애 뉴스 라면 //response => 링크가 리다이렉션인지(사이트 주소가 바뀌는지) 확인하기 위해 response.url 로 만듬
            if "entertain" in response.url:
                # 연예뉴스 사이트 css선택자 사용
                title = soup.select_one(".end_tit")
                content = soup.select_one("#articeBody")
            # 그게 아니고 스포츠 뉴스 url 이면
            elif "sports" in response.url:
                # 스포츠뉴스 사이트 css선택자 사용
                title = soup.select_one("h4.title")
                content = soup.select_one("#newsEndContents")
                #본문 내용안에 불필요한 div, p 삭제
                divs = content.select("div")
                for div in divs:
                    div.decompose()
                paragraphs= content.select("p")
                for p in paragraphs:
                    p.decompose()


            # 그게 아니면
            else:
                # 일반뉴스 사이트 css선택자 사용
                title = soup.select_one(".media_end_head_title")
                content = soup.select_one("#dic_area")
                brs = content.select("br")
                for br in brs:
                    br.decompose()
                brs = content.select("br")
                for br in brs:
                    br.decompose()

            print("=======링크======= \n", url)
            # 공백 제거 .text.strip()
            print("=======제목======= \n", title.text.strip())

            print("=======본문======= \n", content.text.strip())
            num = 1
            # 셀 데이터 추가하기
            ws[f'A{row}'] = url
            ws[f'B{row}'] = title.text.strip()
            ws[f'C{row}'] = content.text.strip()

            #자동 줄바꿈
            ws[f'C{row}'].alignment = Alignment(wrap_text=True)
            row = row + 1
            # time.sleep(0.3)
    page_num = page_num+1

# 엑셀 저장하기
wb.save(f"{keyword}.xlsx")
