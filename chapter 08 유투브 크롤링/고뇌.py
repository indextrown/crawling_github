from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager  #크롬 드라이버 자동 업데이트
import requests
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys  # 엔터키 관련
import time
import openpyxl
import re


# 입력부분
keyword = input("검색어를 입력하시오")

# 1 엑셀 만들기
wb = openpyxl.Workbook()

# 입력한 키워드 명으로 엑셀 시트 생성
ws = wb.create_sheet(keyword)

# 처음 새로운 엑셀파일이 생성될 때 존재하는 "Sheet" 시트 삭제
wb.remove(wb['Sheet'])

# 열 너비 조절
ws.column_dimensions['A'].width = 5.0
ws.column_dimensions['B'].width = 70.0
ws.column_dimensions['C'].width = 10.0
ws.column_dimensions['D'].width = 10.0
ws.column_dimensions['E'].width = 70.0

# 리스트 형태로 맨위의 칸 추가
ws.append(['번호', '제목', '조회수', '날짜', '링크'])


#브라우저 꺼짐 방지
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)

#불필요한 에러 메시지 없애기
chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])

#서비스 객체를 만들어서 service에 저장
service = Service(executable_path=ChromeDriverManager().install())

##크롬 브라우저를 생성
#셀레니움 웹드라이버 => chrome을 만들어냄 서비스는 service 집어넣음
driver = webdriver.Chrome(service=service, options=chrome_options)


#웹페이지 해당 주소 이동
driver.implicitly_wait(5) #웹페이지가 로딩 될때까지 5초는 기다림
driver.maximize_window()  #화면 최대화
driver.get(f"https://www.youtube.com/results?search_query={keyword}")


# 스크롤 전 높이  자바스크립트 실행 가능 함수
before_h = driver.execute_script("return window.scrollY")


# 특정 횟수만큼 스크롤 내리기 1~11이면 1 ~ 3을 의미

for i in range(1, 2):
    driver.find_element(By.CSS_SELECTOR, "body").send_keys(Keys.END)
    time.sleep(1)


soup = BeautifulSoup(driver.page_source, 'lxml')

infos = soup.select("div.text-wrapper")

# 번호
num = 1
for info in infos:
    # 제목
    title = info.select_one("a#video-title").text

    # 조회수
    johae = info.select_one("div#metadata-line > span:nth-child(2)").text
    johae_Numbers = re.sub(r'[^0-9]', '', johae)
    print(johae_Numbers)
    #print(johae)







    # 날짜
    date = info.select_one("div#metadata-line > span:nth-child(3)").text
    #print(date)

    # 링크
    link = info.select_one("a#video-title")
    sub_url = "https://www.youtube.com/" + link.attrs['href']
    print(sub_url)
    ws.append([num, title, johae_Numbers, date, sub_url])
    num +=1



# 엑셀 저장하기
#wb.save(f"./{keyword}.xlsx")
wb.save(f"유투브 크롤링 결과/{keyword}.xlsx")

## 하이퍼링크를 달기 위한 코드 과정
from openpyxl import load_workbook
wb = load_workbook(f"유투브 크롤링 결과/{keyword}.xlsx")

# 시트 활성화
ws = wb.active

# 하이퍼링크 활성화
for i in range(2, ws.max_row+1):
    ws['E' + str(i)].hyperlink = ws['E'+str(i)].value
    ws['E' + str(i)].style = "Hyperlink"
wb.save(f"유투브 크롤링 결과/{keyword}_hyperlink.xlsx")



driver.close()
